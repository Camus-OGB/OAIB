"""Views pour l'app exams (éditions, phases, QCM, sessions d'examen)."""
import io
import json

from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Avg, Count

from rest_framework import generics, permissions, status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from apps.permissions import IsAdmin, IsStudent, ReadOnly
from .models import (
    Edition, Phase, QuestionCategory, Question, QuestionOption,
    Exam, ExamQuestion, ExamSession, ExamAnswer,
)
from .serializers import (
    EditionSerializer, PhaseSerializer,
    QuestionCategorySerializer, QuestionSerializer, QuestionCreateSerializer,
    ExamSerializer, ExamDetailSerializer,
    ExamQuestionPublicSerializer,
    ExamSessionSerializer, ExamAnswerSerializer, SubmitAnswerSerializer,
)


# ──────────────────────────────────────────────
# EDITIONS & PHASES (lecture publique, écriture admin)
# ──────────────────────────────────────────────
class EditionViewSet(viewsets.ModelViewSet):
    queryset = Edition.objects.all()
    serializer_class = EditionSerializer

    def get_permissions(self):
        if self.action in ('list', 'retrieve'):
            return [permissions.AllowAny()]
        return [IsAdmin()]

    @action(detail=False, methods=['get'])
    def active(self, request):
        """Retourne l'édition active."""
        edition = Edition.objects.filter(is_active=True).first()
        if not edition:
            return Response({'detail': 'Aucune édition active.'}, status=404)
        return Response(EditionSerializer(edition).data)


class PhaseViewSet(viewsets.ModelViewSet):
    queryset = Phase.objects.select_related('edition').all()
    serializer_class = PhaseSerializer
    filterset_fields = ['edition', 'status']

    def get_permissions(self):
        if self.action in ('list', 'retrieve'):
            return [permissions.AllowAny()]
        return [IsAdmin()]


# ──────────────────────────────────────────────
# QUESTIONS (admin uniquement)
# ──────────────────────────────────────────────
class QuestionCategoryViewSet(viewsets.ModelViewSet):
    queryset = QuestionCategory.objects.all()
    serializer_class = QuestionCategorySerializer
    permission_classes = [IsAdmin]


class QuestionViewSet(viewsets.ModelViewSet):
    queryset = Question.objects.select_related('category').prefetch_related('options').all()
    permission_classes = [IsAdmin]
    filterset_fields = ['category', 'difficulty', 'is_active']
    search_fields = ['text']
    ordering_fields = ['created_at', 'difficulty', 'usage_count']

    def get_serializer_class(self):
        if self.action in ('create', 'update', 'partial_update'):
            return QuestionCreateSerializer
        return QuestionSerializer

    # ── Export JSON ────────────────────────────────────────────
    @action(detail=False, methods=['get'], url_path='export-json')
    def export_json(self, request):
        """Exporte toutes les questions en JSON."""
        qs = self.filter_queryset(self.get_queryset())
        data = []
        for q in qs:
            data.append({
                'text': q.text,
                'category': q.category.name if q.category else '',
                'difficulty': q.difficulty,
                'points': q.points,
                'time_limit_seconds': q.time_limit_seconds,
                'is_active': q.is_active,
                'options': [
                    {'text': o.text, 'is_correct': o.is_correct, 'order': o.order}
                    for o in q.options.all().order_by('order')
                ],
            })
        response = HttpResponse(
            json.dumps(data, ensure_ascii=False, indent=2),
            content_type='application/json',
        )
        response['Content-Disposition'] = 'attachment; filename="questions_oaib.json"'
        return response

    # ── Export Excel ───────────────────────────────────────────
    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """Exporte toutes les questions en Excel (.xlsx)."""
        qs = self.filter_queryset(self.get_queryset())
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Questions'

        # Styles
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1A535C', end_color='1A535C', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        # Headers
        headers = [
            'Question', 'Catégorie', 'Difficulté', 'Points',
            'Temps (s)', 'Active',
            'Option A', 'Correcte A',
            'Option B', 'Correcte B',
            'Option C', 'Correcte C',
            'Option D', 'Correcte D',
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Data rows
        for row_idx, q in enumerate(qs, 2):
            opts = list(q.options.all().order_by('order'))
            ws.cell(row=row_idx, column=1, value=q.text).border = thin_border
            ws.cell(row=row_idx, column=2, value=q.category.name if q.category else '').border = thin_border
            ws.cell(row=row_idx, column=3, value=q.difficulty).border = thin_border
            ws.cell(row=row_idx, column=4, value=q.points).border = thin_border
            ws.cell(row=row_idx, column=5, value=q.time_limit_seconds).border = thin_border
            ws.cell(row=row_idx, column=6, value='Oui' if q.is_active else 'Non').border = thin_border
            for i in range(4):
                if i < len(opts):
                    ws.cell(row=row_idx, column=7 + i * 2, value=opts[i].text).border = thin_border
                    cell = ws.cell(row=row_idx, column=8 + i * 2, value='Oui' if opts[i].is_correct else 'Non')
                    cell.border = thin_border
                    if opts[i].is_correct:
                        cell.font = Font(bold=True, color='00AA00')

        # Column widths
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 18
        for col_letter in 'CDEF':
            ws.column_dimensions[col_letter].width = 12
        for col_letter in ('G', 'I', 'K', 'M'):
            ws.column_dimensions[col_letter].width = 35
        for col_letter in ('H', 'J', 'L', 'N'):
            ws.column_dimensions[col_letter].width = 12

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="questions_oaib.xlsx"'
        return response

    # ── Import JSON ───────────────────────────────────────────
    @action(detail=False, methods=['post'], url_path='import-json', parser_classes=[MultiPartParser, FormParser])
    def import_json(self, request):
        """Importe des questions depuis un fichier JSON."""
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'Fichier requis.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            data = json.loads(file.read().decode('utf-8'))
        except (json.JSONDecodeError, UnicodeDecodeError):
            return Response({'detail': 'Fichier JSON invalide.'}, status=status.HTTP_400_BAD_REQUEST)

        created, errors = self._import_questions(data)
        return Response({'created': created, 'errors': errors})

    # ── Import Excel ──────────────────────────────────────────
    @action(detail=False, methods=['post'], url_path='import-excel', parser_classes=[MultiPartParser, FormParser])
    def import_excel(self, request):
        """Importe des questions depuis un fichier Excel (.xlsx)."""
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'Fichier requis.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception:
            return Response({'detail': 'Fichier Excel invalide.'}, status=status.HTTP_400_BAD_REQUEST)

        data = []
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or not row[0]:
                continue
            options = []
            for i in range(4):
                opt_text = row[6 + i * 2] if len(row) > 6 + i * 2 else None
                opt_correct = row[7 + i * 2] if len(row) > 7 + i * 2 else None
                if opt_text:
                    options.append({
                        'text': str(opt_text),
                        'is_correct': str(opt_correct).lower() in ('oui', 'true', '1', 'yes'),
                        'order': i + 1,
                    })
            data.append({
                'text': str(row[0]),
                'category': str(row[1]) if row[1] else '',
                'difficulty': str(row[2]) if row[2] else 'medium',
                'points': int(row[3]) if row[3] else 1,
                'time_limit_seconds': int(row[4]) if row[4] else 60,
                'is_active': str(row[5]).lower() in ('oui', 'true', '1', 'yes') if row[5] else True,
                'options': options,
            })

        created, errors = self._import_questions(data)
        return Response({'created': created, 'errors': errors})

    def _import_questions(self, data: list) -> tuple:
        """Crée les questions depuis une liste de dicts. Retourne (created_count, errors_list)."""
        created = 0
        errors = []
        for idx, item in enumerate(data):
            try:
                # Get or create category
                cat_name = item.get('category', '').strip()
                category = None
                if cat_name:
                    category, _ = QuestionCategory.objects.get_or_create(
                        name=cat_name,
                        defaults={'slug': cat_name.lower().replace(' ', '-')},
                    )

                q = Question.objects.create(
                    text=item['text'],
                    category=category,
                    difficulty=item.get('difficulty', 'medium'),
                    points=item.get('points', 1),
                    time_limit_seconds=item.get('time_limit_seconds', 60),
                    is_active=item.get('is_active', True),
                )
                for opt in item.get('options', []):
                    QuestionOption.objects.create(
                        question=q,
                        text=opt['text'],
                        is_correct=opt.get('is_correct', False),
                        order=opt.get('order', 1),
                    )
                created += 1
            except Exception as e:
                errors.append({'index': idx, 'error': str(e)})
        return created, errors


# ──────────────────────────────────────────────
# EXAMENS (admin: CRUD, étudiant: voir les siens)
# ──────────────────────────────────────────────
class ExamViewSet(viewsets.ModelViewSet):
    queryset = Exam.objects.select_related('phase').all()
    filterset_fields = ['phase', 'status']
    ordering_fields = ['start_datetime', 'created_at']

    def get_permissions(self):
        if self.action in ('list', 'retrieve'):
            return [permissions.IsAuthenticated()]
        return [IsAdmin()]

    def get_serializer_class(self):
        if self.action == 'retrieve' and self.request.user.role in ('admin', 'moderator'):
            return ExamDetailSerializer
        return ExamSerializer


# ──────────────────────────────────────────────
# SESSION D'EXAMEN (candidat)
# ──────────────────────────────────────────────
class StartExamView(generics.CreateAPIView):
    """Le candidat démarre un examen."""
    permission_classes = [permissions.IsAuthenticated, IsStudent]

    def create(self, request, exam_id=None):
        try:
            exam = Exam.objects.get(pk=exam_id, status='active')
        except Exam.DoesNotExist:
            return Response(
                {'detail': "Examen introuvable ou pas encore actif."},
                status=status.HTTP_404_NOT_FOUND,
            )

        profile = request.user.candidate_profile
        session, created = ExamSession.objects.get_or_create(
            candidate=profile,
            exam=exam,
            defaults={
                'started_at': timezone.now(),
                'status': ExamSession.Status.IN_PROGRESS,
                'max_score': sum(q.points for q in exam.questions.all()),
            },
        )

        if not created and session.status != ExamSession.Status.NOT_STARTED:
            return Response(
                {'detail': "Vous avez déjà commencé ou terminé cet examen."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Retourner les questions (sans les réponses correctes)
        questions = exam.questions.all()
        if exam.randomize_questions:
            questions = questions.order_by('?')

        return Response({
            'session_id': session.id,
            'exam': ExamSerializer(exam).data,
            'questions': ExamQuestionPublicSerializer(questions, many=True).data,
        })


class SubmitAnswerView(generics.GenericAPIView):
    """Le candidat soumet une réponse."""
    serializer_class = SubmitAnswerSerializer
    permission_classes = [permissions.IsAuthenticated, IsStudent]

    def post(self, request, session_id=None):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            session = ExamSession.objects.get(
                pk=session_id,
                candidate__user=request.user,
                status=ExamSession.Status.IN_PROGRESS,
            )
        except ExamSession.DoesNotExist:
            return Response(
                {'detail': "Session introuvable ou déjà terminée."},
                status=status.HTTP_404_NOT_FOUND,
            )

        question_id = serializer.validated_data['question_id']
        option_id = serializer.validated_data.get('option_id')
        is_flagged = serializer.validated_data.get('is_flagged', False)

        question = Question.objects.get(pk=question_id)
        selected_option = QuestionOption.objects.get(pk=option_id) if option_id else None
        is_correct = selected_option.is_correct if selected_option else False

        answer, _ = ExamAnswer.objects.update_or_create(
            session=session,
            question=question,
            defaults={
                'selected_option': selected_option,
                'is_correct': is_correct,
                'is_flagged': is_flagged,
            },
        )

        return Response(ExamAnswerSerializer(answer).data)


class FinishExamView(generics.GenericAPIView):
    """Le candidat termine un examen."""
    permission_classes = [permissions.IsAuthenticated, IsStudent]

    def post(self, request, session_id=None):
        try:
            session = ExamSession.objects.get(
                pk=session_id,
                candidate__user=request.user,
                status=ExamSession.Status.IN_PROGRESS,
            )
        except ExamSession.DoesNotExist:
            return Response(
                {'detail': "Session introuvable ou déjà terminée."},
                status=status.HTTP_404_NOT_FOUND,
            )

        session.completed_at = timezone.now()
        if session.started_at:
            session.time_spent_seconds = int(
                (session.completed_at - session.started_at).total_seconds()
            )
        session.status = ExamSession.Status.COMPLETED

        # Calcul du score
        answers = session.answers.all()
        session.score = sum(
            a.question.points for a in answers if a.is_correct
        )
        session.percentage = (
            round(session.score / session.max_score * 100, 2)
            if session.max_score > 0 else 0
        )
        session.save()

        return Response(ExamSessionSerializer(session).data)


class MyExamSessionsView(generics.ListAPIView):
    """Le candidat voit toutes ses sessions."""
    serializer_class = ExamSessionSerializer
    permission_classes = [permissions.IsAuthenticated, IsStudent]

    def get_queryset(self):
        return ExamSession.objects.filter(
            candidate__user=self.request.user,
        ).select_related('exam')


# ──────────────────────────────────────────────
# ADMIN — RÉSULTATS & STATISTIQUES
# ──────────────────────────────────────────────
class ExamSessionAdminViewSet(viewsets.ReadOnlyModelViewSet):
    """Toutes les sessions (admin)."""
    queryset = ExamSession.objects.select_related('candidate__user', 'exam').all()
    serializer_class = ExamSessionSerializer
    permission_classes = [IsAdmin]
    filterset_fields = ['exam', 'status', 'candidate']
    ordering_fields = ['percentage', 'started_at', 'score']

    @action(detail=False, methods=['get'])
    def exam_stats(self, request):
        """Statistiques par examen."""
        exam_id = request.query_params.get('exam_id')
        if not exam_id:
            return Response({'detail': 'exam_id requis.'}, status=400)

        sessions = ExamSession.objects.filter(exam_id=exam_id, status__in=['completed', 'evaluated'])
        stats = sessions.aggregate(
            avg_score=Avg('percentage'),
            total=Count('id'),
        )
        passed = sessions.filter(
            percentage__gte=Exam.objects.get(pk=exam_id).passing_score
        ).count()

        return Response({
            'total_sessions': stats['total'],
            'average_score': round(stats['avg_score'] or 0, 2),
            'passed': passed,
            'failed': stats['total'] - passed,
        })
